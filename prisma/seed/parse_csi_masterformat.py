"""
One-shot parser: reads prisma/seed/data/CSI_MasterFormat_2020.xlsx and
emits prisma/seed/csi_masterformat_seed.json with Level 3 entries only.

Level 3 = 6-digit CSI section codes ("03 30 00"). Level 4 cost-code
subsections ("03 30 00.51") are filtered out — not used by spec books.

Run with: sidecar/.venv/Scripts/python prisma/seed/parse_csi_masterformat.py
"""

import json
import os
import re
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parents[2]
SOURCE = REPO_ROOT / "prisma" / "seed" / "data" / "CSI_MasterFormat_2020.xlsx"
OUT = REPO_ROOT / "prisma" / "seed" / "csi_masterformat_seed.json"

# Level 3 pattern: exactly "XX XX XX" — no decimal suffix
LEVEL_3_RE = re.compile(r"^\d{2}\s\d{2}\s\d{2}$")

# Division titles from the CATEGORIES sheet
DIVISION_TITLES = {
    "00": "General Project Requirements",
    "01": "General Requirements",
    "02": "Existing Conditions",
    "03": "Concrete",
    "04": "Masonry",
    "05": "Metals",
    "06": "Wood, Plastics, and Composites",
    "07": "Thermal and Moisture Protection",
    "08": "Openings",
    "09": "Finishes",
    "10": "Specialties",
    "11": "Equipment",
    "12": "Furnishings",
    "13": "Special Construction",
    "14": "Conveying Equipment",
    "21": "Fire Suppression",
    "22": "Plumbing",
    "23": "Heating, Ventilating, and Air Conditioning (HVAC)",
    "25": "Integrated Automation",
    "26": "Electrical",
    "27": "Communications",
    "28": "Electronic Safety and Security",
    "31": "Earthwork",
    "32": "Exterior Improvements",
    "33": "Utilities",
    "34": "Transportation",
    "35": "Waterway and Marine Construction",
    "40": "Process Integration",
    "41": "Material Processing and Handling Equipment",
    "42": "Process Heating, Cooling, and Drying Equipment",
    "43": "Process Gas and Liquid Handling, Purification, and Storage Equipment",
    "44": "Pollution and Waste Control Equipment",
    "45": "Industry-Specific Manufacturing Equipment",
    "46": "Water and Wastewater Equipment",
    "48": "Electrical Power Generation",
}


def main() -> None:
    if not SOURCE.exists():
        raise SystemExit(f"Source file not found: {SOURCE}")

    wb = openpyxl.load_workbook(SOURCE, data_only=True)

    entries: list[dict] = []
    division_counts: dict[str, int] = {}

    for sheet_name in wb.sheetnames:
        if sheet_name == "CATEGORIES":
            continue
        ws = wb[sheet_name]

        # Each sheet: row 1 is header ("Cost-Code" | "Description"), rows 2+ are data
        for row in ws.iter_rows(min_row=2, values_only=True):
            code = row[0]
            title = row[1]
            if not code or not title:
                continue

            code = str(code).strip()
            title = str(title).strip()

            # Keep only Level 3 (6-digit, no decimal)
            if not LEVEL_3_RE.match(code):
                continue

            division = code.split(" ")[0]

            entries.append({
                "csiNumber": code,
                "canonicalTitle": title,
                "division": division,
            })
            division_counts[division] = division_counts.get(division, 0) + 1

    # Sort by CSI number for reproducibility
    entries.sort(key=lambda e: e["csiNumber"])

    # Dedupe by csiNumber (keep first occurrence)
    seen: set[str] = set()
    deduped: list[dict] = []
    for e in entries:
        if e["csiNumber"] in seen:
            continue
        seen.add(e["csiNumber"])
        deduped.append(e)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    with OUT.open("w", encoding="utf-8") as f:
        json.dump(deduped, f, indent=2, ensure_ascii=False)

    print(f"[parse_csi] wrote {len(deduped)} Level 3 entries to {OUT.relative_to(REPO_ROOT)}")
    print(f"[parse_csi] division breakdown:")
    for div in sorted(division_counts.keys()):
        title = DIVISION_TITLES.get(div, "?")
        print(f"   {div} ({title}): {division_counts[div]}")


if __name__ == "__main__":
    main()
